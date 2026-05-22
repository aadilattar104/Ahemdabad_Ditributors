"""
modern_trade_parser.py
======================
Parses Reliance-style Modern Trade Excel reports (Fresh Pik / GoFresh formats).

Supports:
  • FY 25-26 sheet  — columns: store_code@4, store_name@5, article_id@15, sku@16,
                               month-pairs starting @17 (Aug Qty, Aug Sales, ...)
  • FY 26-27 sheet  — columns: store_code@3, store_name@4, article_id@16, sku@17,
                               month-pairs starting @18 (Apr'26 Qty, Apr'26 Sales, ...)
  • SOH sheet       — flat table: site_code@1, site_name@2, article@3, desc@4,
                                  mrp_type@5, soh_qty@6, soh_value@7, map@8

Zero shared code with generic_parser.py. Completely additive.
"""

import io
import re
import openpyxl

# ─── Month name normalisation ──────────────────────────────────────────────────
_MONTH_ABBR = {
    "jan": ("January",  2026),
    "feb": ("February", 2026),
    "mar": ("March",    2026),
    "apr": ("April",    2026),
    "may": ("May",      2026),
    "jun": ("June",     2026),
    "jul": ("July",     2026),
    "aug": ("August",   2025),
    "sep": ("September",2025),
    "oct": ("October",  2025),
    "nov": ("November", 2025),
    "dec": ("December", 2025),
}

def _parse_month_col(col_name: str):
    """
    Parse column headers like:
      "Aug Qty"   → ("August",  2025, "qty")
      "Aug Sales" → ("August",  2025, "sales")
      "Apr'26 Qty "  → ("April",   2026, "qty")
      "Apr'26 Sales" → ("April",   2026, "sales")
    Returns (month_str, year_int, kind) or None if not a month column.
    """
    if not col_name:
        return None
    s = str(col_name).strip()

    # Pattern 1: "Apr'26 Qty" / "Apr'26 Sales"
    m = re.match(r"([A-Za-z]{3})'(\d{2})\s+(Qty|Sales)", s, re.IGNORECASE)
    if m:
        abbr = m.group(1).lower()
        year = 2000 + int(m.group(2))
        kind = m.group(3).lower()
        month_info = _MONTH_ABBR.get(abbr)
        if month_info:
            return (month_info[0], year, kind)

    # Pattern 2: "Aug Qty" / "Sep Sales"
    m = re.match(r"([A-Za-z]{3})\s+(Qty|Sales)$", s, re.IGNORECASE)
    if m:
        abbr = m.group(1).lower()
        kind = m.group(2).lower()
        month_info = _MONTH_ABBR.get(abbr)
        if month_info:
            return (month_info[0], month_info[1], kind)

    return None


def _is_soh_sheet(header_row) -> bool:
    """Detect SOH sheet by looking for 'Site Code' or 'SOH Qty' in the header."""
    if not header_row:
        return False
    for cell in header_row:
        if isinstance(cell, str) and cell.strip().lower() in ("site code", "soh qty"):
            return True
    return False


def _parse_sales_sheet(rows, chain_name: str, upload_id: str, sheet_name: str) -> list:
    """
    Given all rows from a sales sheet (FY 25-26 or FY 26-27), return a list of
    mt_sales_records dicts. Zero-qty AND zero-revenue rows are skipped.
    """
    if len(rows) < 3:
        return []

    header = rows[1]   # row index 1 is the column header row

    # Detect which column layout we're in by checking position of "Parent Store No"
    # FY 25-26: store_code @4, store_name @5, article_id @15, sku_name @16, months @17+
    # FY 26-27: store_code @3, store_name @4, article_id @16, sku_name @17, months @18+
    if str(header[3] or "").strip() == "Parent Store No":
        store_code_col  = 3
        store_name_col  = 4
        article_id_col  = 16
        sku_name_col    = 17
        month_start_col = 18
    else:
        # Default to FY 25-26 layout
        store_code_col  = 4
        store_name_col  = 5
        article_id_col  = 15
        sku_name_col    = 16
        month_start_col = 17

    # Build month column index map: { col_idx: (month, year, kind) }
    month_cols = {}
    for ci, col_name in enumerate(header):
        parsed = _parse_month_col(col_name)
        if parsed:
            month_cols[ci] = parsed   # (month_str, year_int, "qty"/"sales")

    if not month_cols:
        print(f"[MT PARSER] WARNING: no month columns found in sheet '{sheet_name}'")
        return []

    # Group month columns into pairs: {(month, year): {"qty": idx, "sales": idx}}
    month_pair = {}
    for ci, (month_str, year_int, kind) in month_cols.items():
        key = (month_str, year_int)
        if key not in month_pair:
            month_pair[key] = {}
        month_pair[key][kind] = ci

    records = []
    for row in rows[2:]:   # data rows start at index 2
        if row is None:
            continue

        store_code = str(row[store_code_col] or "").strip()
        store_name = str(row[store_name_col] or "").strip()
        article_id = str(row[article_id_col] or "").strip()
        sku_name   = str(row[sku_name_col]   or "").strip()

        if not store_code or not sku_name:
            continue

        for (month_str, year_int), cols in month_pair.items():
            qty_col   = cols.get("qty")
            sales_col = cols.get("sales")

            qty = 0
            if qty_col is not None and row[qty_col] is not None:
                try:
                    qty = int(float(row[qty_col]))
                except (ValueError, TypeError):
                    qty = 0

            revenue = 0.0
            if sales_col is not None and row[sales_col] is not None:
                try:
                    revenue = float(row[sales_col])
                except (ValueError, TypeError):
                    revenue = 0.0

            # Skip completely empty month-SKU-store combos
            if qty == 0 and revenue == 0:
                continue

            records.append({
                "upload_id":  upload_id,
                "chain_name": chain_name,
                "store_code": store_code,
                "store_name": store_name,
                "article_id": article_id,
                "sku_name":   sku_name,
                "qty":        qty,
                "revenue":    round(revenue, 2),
                "month":      month_str,
                "year":       year_int,
            })

    return records


def _parse_soh_sheet(rows, chain_name: str, upload_id: str) -> list:
    """
    Parse the SOH flat table.
    Header is at row index 1: [_, Site Code, Site Name, Article, Article description,
                                MRP Type, SOH Qty, SOH Value, MAP]
    """
    records = []
    for row in rows[2:]:
        if row is None:
            continue
        site_code = str(row[1] or "").strip()
        site_name = str(row[2] or "").strip()
        article   = str(row[3] or "").strip()
        sku_name  = str(row[4] or "").strip()
        soh_qty   = 0
        soh_value = 0.0
        map_price = 0.0

        if not site_code or not sku_name:
            continue

        try:
            soh_qty = int(float(row[6])) if row[6] is not None else 0
        except (ValueError, TypeError):
            soh_qty = 0
        try:
            soh_value = float(row[7]) if row[7] is not None else 0.0
        except (ValueError, TypeError):
            soh_value = 0.0
        try:
            map_price = float(row[8]) if row[8] is not None else 0.0
        except (ValueError, TypeError):
            map_price = 0.0

        records.append({
            "upload_id":  upload_id,
            "chain_name": chain_name,
            "store_code": site_code,
            "store_name": site_name,
            "article_id": article,
            "sku_name":   sku_name,
            "soh_qty":    soh_qty,
            "soh_value":  round(soh_value, 2),
            "map_price":  round(map_price, 2),
        })

    return records


# ─── Public entry point ────────────────────────────────────────────────────────

def run_mt_extraction_pipeline(
    file_bytes: bytes,
    filename: str,
    chain_name: str,
    upload_id: str,
) -> dict:
    """
    Parse a Modern Trade Excel file and return:
      {
        sales_records: [...],   # ready to INSERT into mt_sales_records
        soh_records:   [...],   # ready to INSERT into mt_soh_records
        stores_found:  [...],   # list of store_code strings
        sales_count:   int,
        soh_count:     int,
      }

    This function is completely independent of run_extraction_pipeline().
    It does NOT touch sales_records, uploads, or any distributor tables.
    """
    chain_name = chain_name.strip().upper()

    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True
    )

    all_sales = []
    all_soh   = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        header = rows[1]

        if _is_soh_sheet(header):
            print(f"[MT PARSER] Sheet '{sheet_name}' → SOH")
            soh = _parse_soh_sheet(rows, chain_name, upload_id)
            all_soh.extend(soh)
            print(f"[MT PARSER] SOH records: {len(soh)}")
        else:
            print(f"[MT PARSER] Sheet '{sheet_name}' → Sales")
            sales = _parse_sales_sheet(rows, chain_name, upload_id, sheet_name)
            all_sales.extend(sales)
            print(f"[MT PARSER] Sales records: {len(sales)}")

    stores_found = list({r["store_code"] for r in all_sales + all_soh})

    print(f"[MT PARSER] TOTAL → sales={len(all_sales)} soh={len(all_soh)} stores={stores_found}")

    return {
        "sales_records": all_sales,
        "soh_records":   all_soh,
        "stores_found":  stores_found,
        "sales_count":   len(all_sales),
        "soh_count":     len(all_soh),
    }